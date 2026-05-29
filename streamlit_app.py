"""
# 访谈智析 (InterviewInsight) - Streamlit 网页应用
# 部署到 Streamlit Cloud: https://streamlit.io/cloud
# 需要 OpenAI API Key (配置在 Secrets 中)
"""

import os
import io
import json
import tempfile
import subprocess
import uuid
from datetime import datetime

import streamlit as st
import openai
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============================================================
# 页面配置
# ============================================================
st.set_page_config(
    page_title="访谈智析 · InterviewInsight",
    page_icon="🎙️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# 样式
# ============================================================
st.markdown("""
<style>
    .main-header {
        font-size: 2.2rem;
        font-weight: 700;
        color: #1E3A5F;
        margin-bottom: 0.2rem;
    }
    .sub-header {
        font-size: 1rem;
        color: #6B7280;
        margin-bottom: 1.5rem;
    }
    .stat-card {
        background: #F0F4F8;
        border-radius: 12px;
        padding: 1rem;
        text-align: center;
        border: 1px solid #E2E8F0;
    }
    .stat-number {
        font-size: 1.8rem;
        font-weight: 700;
        color: #1E3A5F;
    }
    .stat-label {
        font-size: 0.8rem;
        color: #6B7280;
    }
    .download-btn {
        display: inline-block;
        padding: 0.5rem 1.5rem;
        background: #1E3A5F;
        color: white;
        border-radius: 8px;
        text-decoration: none;
        font-weight: 600;
    }
    footer { display: none; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# 辅助函数
# ============================================================

def get_openai_client():
    """获取 OpenAI 客户端"""
    api_key = st.secrets.get("OPENAI_API_KEY") or os.getenv("OPENAI_API_KEY")
    if not api_key:
        st.error("⚠️ 未配置 OpenAI API Key。请在 Streamlit Cloud Secrets 中设置 `OPENAI_API_KEY`")
        st.info("🔑 获取 API Key：https://platform.openai.com/api-keys")
        st.stop()
    return openai.OpenAI(api_key=api_key)


def check_ffmpeg():
    """检查 ffmpeg 是否可用"""
    try:
        subprocess.run(["ffmpeg", "-version"], capture_output=True, check=True)
        return True
    except Exception:
        return False


def convert_audio_to_wav(input_path):
    """用 ffmpeg 将音频转为 16kHz mono WAV"""
    output_path = tempfile.mktemp(suffix=".wav")
    cmd = ["ffmpeg", "-y", "-i", input_path,
           "-ar", "16000", "-ac", "1", "-sample_fmt", "s16",
           "-f", "wav", output_path]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        raise RuntimeError(f"ffmpeg 转码失败: {result.stderr[-200:]}")
    return output_path


def transcribe_audio(client, audio_path):
    """调用 OpenAI Whisper API 转写音频"""
    with open(audio_path, "rb") as f:
        transcript = client.audio.transcriptions.create(
            model="whisper-1",
            file=f,
            response_format="text",
            language="zh",
        )
    return transcript.strip()


def llm_clean(client, text):
    """用 LLM 清理语气词"""
    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "你是一个文本清理助手。清理文本中的语气词填充词（嗯、啊、那个、就是说、哦、呃等），保留核心语义。只返回清理后的文本。"},
            {"role": "user", "content": f"清理语气词：{text}"}
        ],
        temperature=0.1,
        max_tokens=500,
    )
    return resp.choices[0].message.content.strip()


def extract_answers(client, full_text, questions):
    """用 LLM 从混合文本中提取受访者回答并匹配问题"""
    q_text = "\n".join([f"{q['question_id']}: {q['question_text']}" for q in questions])

    system_prompt = """你是一个专业的访谈分析助手。以下是一段游戏用户研究访谈转写文本（主持人+受访者混合）。

任务：找出所有受访者的回答，匹配到对应问题，清理语气词。

规则：
1. 只提取受访者说过的原话（或大意），排除主持人说的内容
2. 回答通常很简短，即使一两个词也算
3. 清理语气词（嗯、啊、哦、呃等）
4. 如果回答无法匹配任何问题，用 question_id: "unmatched"

输出格式（纯JSON数组，不要其他文字）：
[
  {"question_id": "Q1", "cleaned_text": "清理后的回答", "score": 0.9, "note": "是/推断/不确定"},
  {"question_id": "unmatched", ...}
]

score: 0.0-1.0，表示匹配置信度
note: "是"确定原话 / "推断"上下文推断 / "不确定""""
    
    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"访谈问题：\n{q_text}\n\n转写全文：\n{full_text}"}
        ],
        temperature=0.1,
        max_tokens=2000,
    )
    
    content = resp.choices[0].message.content
    # 尝试提取 JSON
    try:
        # 找第一个 [ 和最后一个 ]
        start = content.find("[")
        end = content.rfind("]")
        if start >= 0 and end > start:
            content = content[start:end+1]
        return json.loads(content)
    except json.JSONDecodeError:
        return []


def code_answers(client, answers):
    """用 LLM 对回答进行初步编码"""
    if not answers:
        return []
    
    system_prompt = """你是一个质性研究编码专家。对每个受访者回答进行初步编码。

输出JSON数组（纯JSON，不要其他文字）：
[
  {
    "question_id": "Q1",
    "raw_text": "优化不行",
    "theme_code": "产品功能评价",
    "sub_code": "功能缺陷",
    "keywords": "优化,卡顿",
    "sentiment": "负面",
    "coding_note": "用户明确提到优化问题"
  }
]

编码维度：
- theme_code: 一级主题编码（如：产品功能评价/情感态度/使用行为/推荐意愿/竞品对比等）
- sub_code: 二级子编码（如：功能缺陷/满意度/情感连接/使用频率/推荐理由等）
- sentiment: 情感倾向（正面/负面/中性）
- keywords: 关键词（逗号分隔）
- coding_note: 编码依据说明"""
    
    answers_text = json.dumps(answers, ensure_ascii=False, indent=2)
    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"请对以下回答进行编码：\n{answers_text}"}
        ],
        temperature=0.1,
        max_tokens=2000,
    )
    
    content = resp.choices[0].message.content
    try:
        start = content.find("[")
        end = content.rfind("]")
        if start >= 0 and end > start:
            content = content[start:end+1]
        return json.loads(content)
    except json.JSONDecodeError:
        return []


def generate_raw_excel(questions, qa_map, output_path):
    """生成原话版 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "原话版"
    
    # 样式
    hf = Font(bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(vertical="top", wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    uf = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # 标题
    ws.merge_cells('A1:E1')
    ws['A1'].value = f"访谈智析 · 原话版报告（{datetime.now().strftime('%Y-%m-%d %H:%M')}）"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    
    # 表头
    headers = ["问题ID", "问题文本", "受访者原话", "置信度", "说明"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
    
    # 数据
    for i, q in enumerate(questions, 3):
        qid = q["question_id"]
        ws.cell(row=i, column=1, value=qid).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=i, column=1).border = tb
        ws.cell(row=i, column=2, value=q["question_text"]).alignment = ca
        ws.cell(row=i, column=2).border = tb
        
        if qid in qa_map:
            answers = qa_map[qid]
            texts = "; ".join([a["text"] for a in answers])
            scores = "; ".join([str(a.get("confidence", "")) for a in answers])
            notes = "; ".join([a.get("note", "✅ 已匹配") for a in answers])
            ws.cell(row=i, column=3, value=texts).alignment = ca
            ws.cell(row=i, column=4, value=scores).alignment = Alignment(horizontal="center")
            ws.cell(row=i, column=5, value=notes).alignment = ca
        else:
            ws.cell(row=i, column=3, value="").alignment = ca
            ws.cell(row=i, column=4, value="").alignment = Alignment(horizontal="center")
            ws.cell(row=i, column=5, value="⚠️ 未获得直接回答").alignment = ca
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = uf
        
        for col in range(3, 6):
            ws.cell(row=i, column=col).border = tb
        ws.row_dimensions[i].height = 25
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    
    wb.save(output_path)


def generate_coded_excel(questions, qa_map, coded_data, output_path):
    """生成编码版 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "编码版"
    
    hf = Font(bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(vertical="top", wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:G1')
    ws['A1'].value = f"访谈智析 · 初步编码版报告（{datetime.now().strftime('%Y-%m-%d %H:%M')}）"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    
    headers = ["问题ID", "问题文本", "受访者原话", "主题编码", "子编码", "关键词", "情感倾向"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
    
    # 构建编码查找表
    code_map = {}
    for c in coded_data:
        qid = c.get("question_id", "")
        if qid not in code_map:
            code_map[qid] = []
        code_map[qid].append(c)
    
    row_idx = 3
    for q in questions:
        qid = q["question_id"]
        
        if qid in qa_map:
            answers = qa_map[qid]
            codes = code_map.get(qid, [])
            
            for ai, a in enumerate(answers):
                text = a["text"]
                code = codes[ai] if ai < len(codes) else {}
                
                ws.cell(row=row_idx, column=1, value=qid if ai == 0 else "").border = tb
                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="top")
                ws.cell(row=row_idx, column=2, value=q["question_text"] if ai == 0 else "").border = tb
                ws.cell(row=row_idx, column=2).alignment = ca
                ws.cell(row=row_idx, column=3, value=text).border = tb
                ws.cell(row=row_idx, column=3).alignment = ca
                ws.cell(row=row_idx, column=4, value=code.get("theme_code", "")).border = tb
                ws.cell(row=row_idx, column=4).alignment = ca
                ws.cell(row=row_idx, column=5, value=code.get("sub_code", "")).border = tb
                ws.cell(row=row_idx, column=5).alignment = ca
                ws.cell(row=row_idx, column=6, value=code.get("keywords", "")).border = tb
                ws.cell(row=row_idx, column=6).alignment = ca
                ws.cell(row=row_idx, column=7, value=code.get("sentiment", "")).border = tb
                ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center")
                ws.row_dimensions[row_idx].height = 22
                row_idx += 1
        else:
            ws.cell(row=row_idx, column=1, value=qid).border = tb
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=q["question_text"]).border = tb
            ws.cell(row=row_idx, column=2).alignment = ca
            for c in range(3, 8):
                ws.cell(row=row_idx, column=c, value="—").border = tb
                ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center")
            ws.row_dimensions[row_idx].height = 22
            row_idx += 1
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 12
    
    wb.save(output_path)


# ============================================================
# 页面 UI
# ============================================================

# 侧边栏 - API 配置
with st.sidebar:
    st.image("https://img.icons8.com/color/96/microphone.png", width=48)
    st.markdown("### 🎙️ 访谈智析")
    st.markdown("**InterviewInsight**")
    st.divider()
    
    # API Key 输入
    api_key = st.text_input(
        "OpenAI API Key",
        type="password",
        help="在 https://platform.openai.com/api-keys 获取",
        placeholder="sk-..."
    )
    if api_key:
        os.environ["OPENAI_API_KEY"] = api_key
    
    st.divider()
    st.caption("📌 不需要 API Key？")
    st.caption("直接联系我（对话中的AI助手），发送音频+问题列表给你处理")
    st.caption("💡 每次处理约 0.5-1 元（Whisper + GPT-4o-mini）")


# 主页面
col1, col2 = st.columns([3, 2])

with col1:
    st.markdown('<div class="main-header">🎙️ 访谈智析</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">上传访谈录音 + 问题框架，自动输出结构化原话报告与主题编码</div>', unsafe_allow_html=True)

with col2:
    st.markdown("""
    <div style="text-align:right; padding-top:0.5rem;">
        <span style="background:#E8F4FD; padding:0.3rem 1rem; border-radius:20px; font-size:0.85rem;">
        ⚡ 10分钟完成 8 小时工作
        </span>
    </div>
    """, unsafe_allow_html=True)

# 主工作区
tab1, tab2, tab3 = st.tabs(["📤 上传与处理", "📊 报告预览", "ℹ️ 使用说明"])

with tab1:
    row1 = st.columns([1, 1])
    
    with row1[0]:
        st.markdown("### 🎵 上传访谈音频")
        audio_file = st.file_uploader(
            "支持格式：mp3, m4a, wav, ogg",
            type=["mp3", "m4a", "wav", "ogg", "mp4"],
            label_visibility="collapsed"
        )
        if audio_file:
            st.audio(audio_file, format=f"audio/{audio_file.name.split('.')[-1]}")
            st.success(f"✅ 已上传：{audio_file.name}")
    
    with row1[1]:
        st.markdown("### 📋 访谈问题框架")
        input_mode = st.radio("输入方式", ["JSON 编辑", "上传 JSON"], horizontal=True, label_visibility="collapsed")
        
        default_questions = json.dumps([
            {"question_id": "Q1", "question_text": "最近一周是否玩过？主要玩什么内容/模式？"},
            {"question_id": "Q2", "question_text": "有什么感觉？为什么玩这个？"},
            {"question_id": "Q3", "question_text": "自己玩还是和朋友玩？"},
            {"question_id": "Q4", "question_text": "还玩什么其他游戏？"},
            {"question_id": "Q5", "question_text": "满意度打几分？（1-5分）"},
            {"question_id": "Q6", "question_text": "满意的部分有哪些？"},
            {"question_id": "Q7", "question_text": "最兴奋/开心的事是什么？"},
            {"question_id": "Q8", "question_text": "最吸引你坚持玩的原因？"},
            {"question_id": "Q9", "question_text": "不满意的地方？"},
            {"question_id": "Q10", "question_text": "画面美术音效？"},
            {"question_id": "Q11", "question_text": "关卡难度？"},
            {"question_id": "Q12", "question_text": "推荐意愿？（0-10分）"},
            {"question_id": "Q13", "question_text": "不太想推荐的原因？"},
            {"question_id": "Q14", "question_text": "因画面好推荐过其他游戏吗？"},
            {"question_id": "Q15", "question_text": "推荐过我们的游戏吗？"},
            {"question_id": "Q16", "question_text": "其他建议？"},
        ], ensure_ascii=False, indent=2)
        
        if input_mode == "JSON 编辑":
            questions_json = st.text_area(
                "编辑问题列表（JSON 格式）",
                value=default_questions,
                height=200,
                label_visibility="collapsed"
            )
        else:
            questions_file = st.file_uploader("上传 JSON 文件", type=["json"], label_visibility="collapsed")
            if questions_file:
                questions_json = questions_file.read().decode("utf-8")
            else:
                questions_json = default_questions
    
    # 开始处理
    st.divider()
    
    col_a, col_b, col_c = st.columns([1, 1, 1])
    with col_b:
        process_btn = st.button(
            "🚀 开始处理",
            type="primary",
            use_container_width=True,
            disabled=not audio_file,
        )
    
    if not audio_file:
        st.info("👆 请先上传音频文件")
    
    # 处理逻辑
    if process_btn and audio_file:
        # 解析问题
        try:
            questions = json.loads(questions_json)
            if not isinstance(questions, list):
                st.error("❌ 问题列表必须是 JSON 数组格式")
                st.stop()
        except json.JSONDecodeError as e:
            st.error(f"❌ JSON 格式错误: {e}")
            st.stop()
        
        progress_bar = st.progress(0, text="初始化...")
        status_area = st.empty()
        
        try:
            # Step 1: 检查 ffmpeg
            status_area.info("🔍 检查环境...")
            progress_bar.progress(5, text="检查 ffmpeg...")
            if not check_ffmpeg():
                st.error("❌ ffmpeg 未安装，请安装后重试：`apt install ffmpeg`")
                st.stop()
            
            # Step 2: 获取 OpenAI Client
            client = get_openai_client()
            
            # Step 3: 保存上传的音频
            status_area.info("📥 保存音频文件...")
            progress_bar.progress(10, text="保存音频...")
            suffix = f".{audio_file.name.split('.')[-1]}"
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
                f.write(audio_file.getbuffer())
                audio_path = f.name
            
            # Step 4: 转码为 WAV
            status_area.info("🔄 音频转码中...")
            progress_bar.progress(20, text="ffmpeg 转码中...")
            wav_path = convert_audio_to_wav(audio_path)
            os.unlink(audio_path)
            
            # Step 5: ASR 转写
            status_area.info("🎤 语音转写中（约1-3分钟）...")
            progress_bar.progress(35, text="Whisper 转写中...")
            full_text = transcribe_audio(client, wav_path)
            os.unlink(wav_path)
            
            status_area.success(f"✅ 转写完成：{len(full_text)} 字")
            with st.expander("📝 查看转写全文"):
                st.write(full_text)
            
            # Step 6: 提取回答
            status_area.info("🔍 提取受访者回答并匹配问题...")
            progress_bar.progress(60, text="LLM 分析中...")
            extracted = extract_answers(client, full_text, questions)
            
            # 构建 qa_map
            qa_map = {}
            for item in extracted:
                qid = item.get("question_id", "")
                if qid == "unmatched" or not qid:
                    continue
                text = item.get("cleaned_text", "")
                if not text:
                    continue
                if qid not in qa_map:
                    qa_map[qid] = []
                qa_map[qid].append({
                    "text": text,
                    "confidence": item.get("score", 0.5),
                    "note": item.get("note", "✅ 已提取"),
                })
            
            # 统计
            answered = len(qa_map)
            total_q = len(questions)
            total_answers = sum(len(v) for v in qa_map.values())
            
            progress_bar.progress(75, text="生成报告中...")
            
            # Step 7: 编码
            status_area.info("🏷️ 正在进行主题编码...")
            progress_bar.progress(80, text="LLM 编码中...")
            coded_data = code_answers(client, extracted)
            
            # Step 8: 生成 Excel
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            raw_path = f"/tmp/interview_raw_{timestamp}.xlsx"
            coded_path = f"/tmp/interview_coded_{timestamp}.xlsx"
            
            generate_raw_excel(questions, qa_map, raw_path)
            generate_coded_excel(questions, qa_map, coded_data, coded_path)
            
            progress_bar.progress(100, text="✅ 完成！")
            status_area.success("🎉 处理完成！")
            
            # 展示结果
            st.divider()
            st.markdown("### 📊 处理结果")
            
            # 统计卡片
            mc1, mc2, mc3, mc4 = st.columns(4)
            with mc1:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number">{total_q}</div>
                    <div class="stat-label">问题总数</div>
                </div>
                """, unsafe_allow_html=True)
            with mc2:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number">{answered}</div>
                    <div class="stat-label">已回答</div>
                </div>
                """, unsafe_allow_html=True)
            with mc3:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number">{total_answers}</div>
                    <div class="stat-label">回答条数</div>
                </div>
                """, unsafe_allow_html=True)
            with mc4:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number">{len(full_text)}</div>
                    <div class="stat-label">转写字数</div>
                </div>
                """, unsafe_allow_html=True)
            
            # 下载按钮
            st.divider()
            st.markdown("### 📥 下载报告")
            dcol1, dcol2 = st.columns(2)
            
            with dcol1:
                with open(raw_path, "rb") as f:
                    st.download_button(
                        label="📄 下载原话版 (Excel)",
                        data=f,
                        file_name=f"访谈原话版_{timestamp}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
            
            with dcol2:
                with open(coded_path, "rb") as f:
                    st.download_button(
                        label="🏷️ 下载编码版 (Excel)",
                        data=f,
                        file_name=f"访谈编码版_{timestamp}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
            
            # 回答预览
            with st.expander("📋 查看匹配到的回答", expanded=True):
                for q in questions:
                    qid = q["question_id"]
                    if qid in qa_map:
                        answers = qa_map[qid]
                        for a in answers:
                            confidence_color = "🟢" if a["confidence"] >= 0.8 else "🟡" if a["confidence"] >= 0.6 else "🟠"
                            st.markdown(f"""
                            **{qid}** {q["question_text"]}
                            > 💬 {a["text"]}  {confidence_color} 置信度：{a["confidence"]:.0%}
                            """)
                    else:
                        st.markdown(f"""
                        **{qid}** {q["question_text"]}
                        > ⚠️ 未获得直接回答
                        """)
            
            # 清理临时文件
            try:
                os.unlink(raw_path)
                os.unlink(coded_path)
            except:
                pass
            
        except Exception as e:
            progress_bar.empty()
            status_area.error(f"❌ 处理出错：{str(e)}")
            st.exception(e)

with tab2:
    st.info("📊 处理完成后，这里会显示报告预览和统计数据")
    st.markdown("""
    **报告说明：**
    
    **原话版（Raw）**
    | 问题ID | 问题文本 | 受访者原话 | 置信度 | 说明 |
    |--------|---------|-----------|--------|------|
    | Q1 | 最近一周是否玩过？ | | | ⚠️ 未获得直接回答 |
    | Q5 | 满意度打几分？ | 给3分 | 0.70 | ✅ 已提取 |
    
    **编码版（Coded）**
    | 问题ID | 问题文本 | 受访者原话 | 主题编码 | 子编码 | 关键词 | 情感 |
    |--------|---------|-----------|---------|-------|--------|------|
    | Q9 | 不满意的地方？ | 优化不行 | 产品功能评价 | 功能缺陷 | 优化,卡顿 | 负面 |
    
    先上传音频和问题列表，点击"开始处理"即可。
    """)

with tab3:
    st.markdown("""
    ## ℹ️ 使用说明
    
    ### 🚀 快速开始
    1. **上传音频**：点击左侧上传区，选择访谈录音文件
    2. **输入问题**：编辑或上传访谈框架 JSON
    3. **点击处理**：等待1-3分钟
    4. **下载报告**：获得原话版 + 编码版两个 Excel
    
    ### 🎯 适用场景
    - 用户研究访谈整理
    - 心理学/教育学质性研究
    - 市场调研数据分析
    - 产品体验反馈梳理
    
    ### ⚠️ 注意事项
    - **单声道录音**：如果主持人和受访者声音混合，提取准确率约50-70%
    - **双声道录音**：建议使用双声道录音（最佳）
    - 音频时长建议不超过30分钟
    - 文件大小建议不超过50MB
    
    ### 💰 费用估算
    | 服务 | 10分钟音频 | 30分钟音频 |
    |------|-----------|-----------|
    | Whisper 转写 | ~$0.06 | ~$0.18 |
    | GPT-4o-mini 分析 | ~$0.01 | ~$0.03 |
    | **合计** | **~0.5元** | **~1.5元** |
    
    ### 🔑 部署到自己账户
    1. 注册 [Streamlit Cloud](https://streamlit.io/cloud)
    2. Fork 这个项目到 GitHub
    3. 在 Streamlit Cloud 中连接仓库
    4. 在 Secrets 中配置 `OPENAI_API_KEY`
    5. 部署完成！
    """)

# 底部
st.divider()
st.caption("🎙️ 访谈智析 InterviewInsight · 基于 OpenAI Whisper + GPT-4o-mini · 由 AI Agent 搭建专家驱动")