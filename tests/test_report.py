"""Tests for Report generation module."""
import sys
import os
import tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "streamlit_app"))

from utils.memory import InterviewMemory
from utils.report import generate_raw_report


class TestReportGeneration:
    """Test Excel report generation"""

    def test_generate_raw_report_basic(self):
        """Test basic raw report generation with sample data"""
        questions = [
            {"question_id": "Q1", "question_text": "你平时玩什么游戏？"},
            {"question_id": "Q2", "question_text": "每天玩多久？"},
        ]
        mem = InterviewMemory()
        mem.store_response("Q1", "我喜欢玩原神", confidence=0.95)
        mem.store_response("Q2", "大概2小时", confidence=0.85)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            output_path = f.name

        try:
            result_path = generate_raw_report(questions, mem, output_path)
            assert os.path.exists(result_path)
            assert result_path == output_path
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_generate_raw_report_empty_answers(self):
        """Test report with some unanswered questions"""
        questions = [
            {"question_id": "Q1", "question_text": "问题1"},
            {"question_id": "Q2", "question_text": "问题2"},
            {"question_id": "Q3", "question_text": "问题3"},
        ]
        mem = InterviewMemory()
        mem.store_response("Q1", "回答1", confidence=0.9)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            output_path = f.name

        try:
            result_path = generate_raw_report(questions, mem, output_path)
            assert os.path.exists(result_path)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_generate_raw_report_no_responses(self):
        """Test report with no responses at all"""
        questions = [
            {"question_id": "Q1", "question_text": "问题1"},
        ]
        mem = InterviewMemory()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            output_path = f.name

        try:
            result_path = generate_raw_report(questions, mem, output_path)
            assert os.path.exists(result_path)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_generate_raw_report_excel_structure(self):
        """Test that the generated Excel has correct structure"""
        from openpyxl import load_workbook

        questions = [
            {"question_id": "Q1", "question_text": "你平时玩什么游戏？"},
        ]
        mem = InterviewMemory()
        mem.store_response("Q1", "我喜欢玩原神", confidence=0.95)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            output_path = f.name

        try:
            generate_raw_report(questions, mem, output_path)
            wb = load_workbook(output_path)
            ws = wb.active
            assert ws.title == "原话版报告"
            # Check headers
            assert ws.cell(1, 1).value == "问题ID"
            assert ws.cell(1, 2).value == "问题文本"
            assert ws.cell(1, 3).value == "受访者原话"
            # Check data
            assert ws.cell(2, 1).value == "Q1"
            assert ws.cell(2, 2).value == "你平时玩什么游戏？"
            assert ws.cell(2, 3).value == "我喜欢玩原神"
            wb.close()
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)