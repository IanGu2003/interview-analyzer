"""Report generation for interview analysis - produces dual-version Excel"""
import os
import re
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from . import llm_utils
from . import memory as mem


def generate_raw_report(
    questions: list[dict],
    interview_memory: mem.InterviewMemory,
    output_path: str,
) -> str:
    """Generate 'raw version' report: all questions listed, responses filled

    Returns output file path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "原话版报告"

    # Styles
    header_font = Font(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Headers
    headers = ["问题ID", "问题文本", "受访者原话", "置信度", "状态"]
    col_widths = [10, 40, 50, 10, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Data rows
    row = 2
    responses = interview_memory.get_all_responses()

    for q in questions:
        qid = q["question_id"]
        qtext = q["question_text"]
        q_responses = responses.get(qid, [])

        if q_responses:
            first = True
            for r in q_responses:
                ws.cell(row=row, column=1, value=qid if first else "").border = thin_border
                ws.cell(row=row, column=2, value=qtext if first else "").border = thin_border
                ws.cell(row=row, column=3, value=r.get("text", "")).border = thin_border
                ws.cell(row=row, column=3).alignment = cell_align
                ws.cell(row=row, column=4, value=r.get("confidence", "")).border = thin_border
                ws.cell(row=row, column=5, value="已匹配").border = thin_border
                for c in range(1, 6):
                    ws.cell(row=row, column=c).alignment = cell_align
                row += 1
                first = False
        else:
            # Unanswered question - yellow highlight
            ws.cell(row=row, column=1, value=qid).border = thin_border
            ws.cell(row=row, column=2, value=qtext).border = thin_border
            ws.cell(row=row, column=3, value="（该问题未获得直接回答）").border = thin_border
            ws.cell(row=row, column=3).alignment = cell_align
            ws.cell(row=row, column=5, value="未获取回答").border = thin_border
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = yellow_fill
                ws.cell(row=row, column=c).alignment = cell_align
            row += 1

    # Add unmatched section
    if interview_memory.unmatched:
        row += 1
        ws.cell(row=row, column=1, value="【未匹配内容】").font = Font(bold=True, color="FF0000")
        row += 1
        for u in interview_memory.unmatched:
            ws.cell(row=row, column=3, value=u.get("text", "")).border = thin_border
            ws.cell(row=row, column=3).alignment = cell_align
            row += 1

    wb.save(output_path)
    return output_path


def generate_coded_report(
    questions: list[dict],
    interview_memory: mem.InterviewMemory,
    output_path: str,
    client=None,
    model: str = "gpt-4o",
) -> str:
    """Generate 'coded version' report with theme codes, keywords, sentiment

    If client is None, codes will be skipped and marked as "待编码"
    Returns output file path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "初步编码版报告"

    # Styles
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers for coded version
    headers = ["问题ID", "问题文本", "受访者回答", "主题编码", "子编码", "关键词", "情感倾向", "编码备注"]
    col_widths = [10, 35, 40, 15, 15, 25, 12, 30]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Data rows
    row = 2
    responses = interview_memory.get_all_responses()

    for q in questions:
        qid = q["question_id"]
        qtext = q["question_text"]
        q_responses = responses.get(qid, [])

        if q_responses:
            first = True
            for r in q_responses:
                answer_text = r.get("text", "")

                # LLM coding
                if client and answer_text:
                    coded = llm_utils.code_response(
                        client, qtext, answer_text, model=model
                    )
                else:
                    coded = {
                        "theme_code": "待编码",
                        "sub_code": "待编码",
                        "keywords": [],
                        "sentiment": "待判断",
                        "memo": "未接入LLM，请人工编码",
                    }

                ws.cell(row=row, column=1, value=qid if first else "").border = thin_border
                ws.cell(row=row, column=2, value=qtext if first else "").border = thin_border
                ws.cell(row=row, column=3, value=answer_text).border = thin_border
                ws.cell(row=row, column=3).alignment = cell_align
                ws.cell(row=row, column=4, value=coded.get("theme_code", "")).border = thin_border
                ws.cell(row=row, column=5, value=coded.get("sub_code", "")).border = thin_border
                ws.cell(row=row, column=6, value=", ".join(coded.get("keywords", []))).border = thin_border
                ws.cell(row=row, column=7, value=coded.get("sentiment", "")).border = thin_border
                ws.cell(row=row, column=8, value=coded.get("memo", "")).border = thin_border
                ws.cell(row=row, column=8).alignment = cell_align

                for c in range(1, 9):
                    ws.cell(row=row, column=c).alignment = cell_align
                row += 1
                first = False

    wb.save(output_path)
    return output_path


def generate_dual_reports(
    questions: list[dict],
    interview_memory: mem.InterviewMemory,
    output_prefix: str,
    client=None,
    model: str = "gpt-4o",
) -> tuple[str, str]:
    """Generate both raw and coded version reports

    Returns (raw_path, coded_path)
    """
    raw_path = f"{output_prefix}_原话版.xlsx"
    coded_path = f"{output_prefix}_初步编码版.xlsx"

    generate_raw_report(questions, interview_memory, raw_path)
    generate_coded_report(questions, interview_memory, coded_path, client, model)

    return raw_path, coded_path